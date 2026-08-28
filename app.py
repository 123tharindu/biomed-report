import datetime
import io
import os
import base64
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageOps, ImageEnhance

from google import genai
import gspread
from google.oauth2.service_account import Credentials
import requests

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import (
    Image as RLImage,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
    HRFlowable,
)
import streamlit as st

# ==========================================
# 1. PAGE CONFIGURATION & EXECUTIVE CSS
# ==========================================
st.set_page_config(
    page_title="Biomed International - AI Lap Scan Portal",
    page_icon="🏥",
    layout="wide",
    initial_sidebar_state="expanded",
)

def get_local_logo_base64(file_path="bmi_logo.png"):
    if os.path.exists(file_path):
        try:
            with open(file_path, "rb") as f:
                encoded = base64.b64encode(f.read()).decode()
                return f"data:image/png;base64,{encoded}"
        except Exception:
            pass
    return "https://via.placeholder.com/120x60.png?text=BMI+Logo"

LOGO_SRC = get_local_logo_base64("bmi_logo.png")

if os.path.exists("bmi_logo.png"):
    st.logo("bmi_logo.png", icon_image="bmi_logo.png")

st.markdown(
    """
<style>
    @import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700;800&display=swap');
    
    html, body, [class*="css"] {
        font-family: 'Plus Jakarta Sans', -apple-system, BlinkMacSystemFont, sans-serif;
    }

    .main .block-container {
        padding-top: 1.5rem !important;
        padding-bottom: 3rem !important;
        max-width: 1200px;
    }
    
    .main { 
        background-color: #F8FAFC; 
    }

    div[data-baseweb="popover"] {
        max-height: 250px !important;
        z-index: 999999 !important;
    }
    div[data-baseweb="popover"] > div {
        max-height: 250px !important;
        overflow-y: auto !important;
    }

    .brand-header {
        background: linear-gradient(135deg, #0F172A 0%, #1E293B 50%, #1E3A8A 100%);
        padding: 20px 24px; 
        border-radius: 16px; 
        color: white;
        box-shadow: 0 10px 25px -5px rgba(15, 23, 42, 0.25); 
        margin-bottom: 24px;
        border: 1px solid rgba(255, 255, 255, 0.1);
    }
    .brand-header h1 { 
        color: #FFFFFF !important; 
        font-size: 20px !important; 
        font-weight: 800 !important; 
        margin: 0 !important; 
        letter-spacing: -0.5px;
    }
    .brand-header p { 
        color: #94A3B8 !important; 
        font-size: 11px !important; 
        margin-top: 4px !important; 
        font-weight: 600; 
        letter-spacing: 0.5px;
        text-transform: uppercase;
    }
    .status-badge {
        background: rgba(34, 197, 94, 0.15); 
        color: #4ADE80;
        border: 1px solid rgba(74, 222, 128, 0.3);
        padding: 4px 10px; 
        border-radius: 20px; 
        font-size: 11px;
        font-weight: 700; 
        display: inline-flex;
        align-items: center;
        gap: 6px;
    }

    .instrument-card {
        background: #FFFFFF; 
        border: 1px solid #E2E8F0;
        border-radius: 14px; 
        padding: 20px; 
        margin-bottom: 20px;
        box-shadow: 0 4px 12px -2px rgba(0, 0, 0, 0.03);
        transition: all 0.2s ease;
    }
    .instrument-card:hover {
        border-color: #CBD5E1;
        box-shadow: 0 8px 16px -4px rgba(0, 0, 0, 0.06);
    }
    
    .section-title { 
        font-size: 16px; 
        font-weight: 800; 
        color: #0F172A; 
        display: flex;
        align-items: center;
        gap: 8px;
        border-bottom: 2px solid #E2E8F0; 
        padding-bottom: 10px; 
        margin-bottom: 16px; 
    }

    .stButton>button {
        border-radius: 10px !important;
        font-weight: 700 !important;
        transition: all 0.2s ease !important;
    }
    .stButton>button[kind="primary"] {
        background: linear-gradient(135deg, #1E3A8A 0%, #0F172A 100%) !important;
        color: white !important; 
        border: none !important;
        box-shadow: 0 4px 12px rgba(30, 58, 138, 0.25) !important;
    }
    .stButton>button[kind="primary"]:hover {
        transform: translateY(-1px);
        box-shadow: 0 6px 16px rgba(30, 58, 138, 0.35) !important;
    }
</style>
""",
    unsafe_allow_html=True,
)

# ==========================================
# 2. DATA LISTS & CATALOG SETUP
# ==========================================
GEMINI_API_KEY = st.secrets.get("GEMINI_API_KEY", os.environ.get("GEMINI_API_KEY", ""))
client = genai.Client(api_key=GEMINI_API_KEY) if GEMINI_API_KEY else None

SL_HOSPITALS = [
    "-- Select Hospital / Institute --",
    "National Hospital of Sri Lanka (NHSL Colombo)",
    "National Hospital Kandy",
    "National Hospital Galle (Karapitiya)",
    "Colombo South Teaching Hospital (Kalubowila)",
    "Colombo North Teaching Hospital (Ragama)",
    "Sri Jayewardenepura General Hospital (SJGH)",
    "Lady Ridgeway Hospital for Children (LRH)",
    "Apeksha Hospital Maharagama",
    "Castle Street Hospital for Women",
    "De Soysa Hospital for Women (DMH)",
    "District General Hospital Hambantota",
    "District General Hospital Matara",
    "District General Hospital Chilaw",
    "District General Hospital Gampaha",
    "District General Hospital Polonnaruwa",
    "District General Hospital Matale",
    "District General Hospital Trincomalee",
    "Teaching Hospital Peradeniya",
    "Teaching Hospital Kurunegala",
    "Teaching Hospital Anuradhapura",
    "Teaching Hospital Badulla",
    "Teaching Hospital Ratnapura",
    "Teaching Hospital Kalutara",
    "Teaching Hospital Jaffna",
    "Teaching Hospital Batticaloa",
    "Teaching Hospital Kuliyapitiya",
    "Teaching Hospital Kegalle",
    "Base Hospital Tangalle",
    "Base Hospital Kamburupitiya",
    "Base Hospital Elpitiya",
    "Base Hospital Balapitiya",
    "Base Hospital Puttalam",
    "Base Hospital Dambulla",
    "Base Hospital Embilipitiya",
    "General Sir John Kotelawala Defence University Hospital (KDU)",
    "Colombo Army Hospital",
    "Sri Lanka Navy General Hospital Welisara",
    "Asiri Surgical Hospital",
    "Asiri Central Hospital",
    "Nawaloka Hospital Colombo",
    "Lanka Hospitals Colombo",
    "Durdans Hospital Colombo",
    "Kings Hospital Colombo",
    "Ruhunu Hospital Galle",
    "Northern Central Hospital Jaffna",
    "Other (Type manually)"
]

DAMAGE_SUGGESTIONS = [
    "-- Select Detailed Technical Damage --",
    "Sealing Cap Damage: Silicone sealing element is torn/damaged. High risk of pneumoperitoneum gas leakage during insufflation.",
    "Insulation Damage: Insulation layer cracked/peeled near the shaft tip. High risk of stray electrical current leaks (HF insulation failure).",
    "Shaft Insulation Micro-Cracks: Flaking detected along middle shaft. High risk of unwanted tissue burns during HF activation.",
    "Shaft Deformation: Outer shaft tube is visibly bent/misaligned, causing severe internal friction and restricting jaw movement.",
    "Jaw Alignment Failure: Working jaws are misaligned with worn-out gripping teeth. Instrument fails to hold tissue securely.",
    "Ratchet Lock Failure: Handle locking mechanism/ratchet teeth worn out. Instrument fails to hold position under tension.",
    "Scissor Blade Bluntness: Scissor blades show heavy dullness and burrs along the cutting edge. Fails clean cutting.",
    "HF Connector Damage: Monopolar/Bipolar terminal pin bent or corroded. Poor electrical contact during electrosurgery.",
    "Trocar Stopcock Leak: Gas valve/stopcock lever worn out and leaking. Cannot maintain stable intra-abdominal pressure.",
    "Corrosion & Pitting: Severe pitting corrosion and rust stains observed near joints due to chemical sterilization.",
    "Pass Inspection: Instrument in optimal condition. No physical defect or operational damage observed."
]

EXCEL_FILE = "Full Laparoscopy Articles Updated master file 07.07.2026.xlsx"

@st.cache_data
def load_catalog(file_path):
    if os.path.exists(file_path):
        try:
            import pandas as pd
            df = pd.read_excel(file_path)
            df.columns = [str(col).strip() for col in df.columns]
            art_col, desc_col = df.columns[0], df.columns[1] if len(df.columns) > 1 else df.columns[0]
            df = df.dropna(subset=[art_col])
            return dict(zip(df[art_col].astype(str).str.strip(), df[desc_col].astype(str).str.strip()))
        except Exception:
            pass
            
    return {
        "BB365R": "Scissors Curved 17mm",
        "BB074R": "Forceps Dissecting",
        "BC051R": "Needle Holder",
        "EK087P": "Sealing Cap"
    }

catalog_dict = load_catalog(EXCEL_FILE)
article_options = sorted(list(catalog_dict.keys()))

def process_and_compress_image(image_file, max_size=(1200, 1200)):
    img = Image.open(image_file)
    img = ImageOps.exif_transpose(img)
    img.thumbnail(max_size, Image.Resampling.LANCZOS)
    if img.mode in ("RGBA", "P"):
        img = img.convert("RGB")
    
    enhancer = ImageEnhance.Contrast(img)
    img = enhancer.enhance(1.25)
    enhancer = ImageEnhance.Sharpness(img)
    img = enhancer.enhance(1.45)
    enhancer = ImageEnhance.Brightness(img)
    img = enhancer.enhance(1.05)
    return img

def analyze_damage_with_ai(image_file, item_name):
    if not client:
        return "API Key not configured properly.", "OK"
    try:
        compressed_img = process_and_compress_image(image_file, max_size=(800, 800))
        prompt = f"Examine surgical instrument '{item_name}' for damage. Line 1: Technical damage (Max 20 words). Line 2: Recommendation (Replace/Repair/Service/OK)."
        
        response = client.models.generate_content(
            model='gemini-2.0-flash',
            contents=[compressed_img, prompt]
        )
        lines = [line.strip() for line in response.text.strip().split('\n') if line.strip()]
        return lines[0] if len(lines) > 0 else "Inspected", lines[1] if len(lines) > 1 else "Service"
    except Exception as e:
        return f"AI Error: {str(e)}", "Service"

# 🔄 GOOGLE SHEET SYNC FUNCTION (WITH ERROR PRINTING)
def sync_to_google_sheet(instruments_data, meta_data):
    webhook_url = st.secrets.get("WEBHOOK_URL", "")
    if webhook_url:
        try:
            payload = []
            for item in instruments_data:
                payload.append({
                    "report_no": meta_data.get("report_no"),
                    "date": meta_data.get("date"),
                    "hospital": meta_data.get("hospital"),
                    "engineer": meta_data.get("engineer"),
                    "article_num": item.get("art_no"),
                    "description": item.get("name"),
                    "sr_number": item.get("sr_no"),
                    "damage": item.get("damage")
                })
            requests.post(webhook_url, json=payload, timeout=10)
            return True, "Webhook Success"
        except Exception as e:
            pass

    try:
        scopes = ["https://www.googleapis.com/auth/spreadsheets"]
        if "gcp_service_account" in st.secrets:
            creds = Credentials.from_service_account_info(dict(st.secrets["gcp_service_account"]), scopes=scopes)
        elif os.path.exists("credentials.json"):
            creds = Credentials.from_service_account_file("credentials.json", scopes=scopes)
        else:
            return False, "Neither st.secrets['gcp_service_account'] nor credentials.json was found!"

        client_gs = gspread.authorize(creds)
        sheet = client_gs.open("Biomed Lap Inspection Summary").sheet1
        
        rows_to_insert = []
        for item in instruments_data:
            rows_to_insert.append([
                meta_data.get("report_no"),     # A: Report No
                meta_data.get("date"),          # B: Date
                meta_data.get("hospital"),      # C: Hospital
                meta_data.get("engineer"),      # D: Inspection Engineer
                item.get("art_no"),             # E: Instruments Article num
                item.get("name"),               # F: Description
                item.get("sr_no", ""),          # G: Machine Compatible/SR Number
                item.get("damage")              # H: Details of Damage
            ])
            
        sheet.append_rows(rows_to_insert)
        return True, "Success"
    except Exception as e:
        return False, str(e)

def generate_professional_excel(instruments_data, hospital_name, engineer_name, report_no, date_str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inspection Summary"
    ws.views.sheetView[0].showGridLines = True

    NAVY_HEADER = "0F172A"
    ICE_BLUE = "F8FAFC"
    WHITE = "FFFFFF"
    BORDER_COLOR = "E2E8F0"
    TEXT_MAIN = "1E293B"

    font_title = Font(name="Segoe UI", size=12, bold=True, color="FFFFFF")
    font_header = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=9, color=TEXT_MAIN)
    font_bold = Font(name="Segoe UI", size=9, bold=True, color=TEXT_MAIN)

    fill_navy = PatternFill(start_color=NAVY_HEADER, end_color=NAVY_HEADER, fill_type="solid")
    fill_zebra = PatternFill(start_color=ICE_BLUE, end_color=ICE_BLUE, fill_type="solid")
    fill_white = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin_border = Side(border_style="thin", color=BORDER_COLOR)
    cell_border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

    ws.merge_cells("A1:F1")
    ws["A1"] = "BIOMED INTERNATIONAL (PVT) LTD — TECHNICAL INSPECTION REPORT"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_navy
    ws["A1"].alignment = align_center
    ws.row_dimensions[1].height = 30

    meta_info = [
        ("Customer / Hospital:", hospital_name, "Inspection Date:", date_str),
        ("Engineer Name:", engineer_name, "Report Ref No:", report_no)
    ]

    for r_idx, row in enumerate(meta_info, start=3):
        ws.cell(row=r_idx, column=1, value=row[0]).font = font_bold
        ws.cell(row=r_idx, column=2, value=row[1]).font = font_data
        ws.cell(row=r_idx, column=4, value=row[2]).font = font_bold
        ws.cell(row=r_idx, column=5, value=row[3]).font = font_data
        ws.row_dimensions[r_idx].height = 20

    headers = ["#", "ARTICLE NO", "INSTRUMENT DESCRIPTION", "TECHNICAL DAMAGE DETAILS", "RECOMMENDATION", "STATUS"]
    header_row = 6
    ws.row_dimensions[header_row].height = 25

    for c_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c_idx, value=header)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = cell_border

    start_data_row = 7
    for idx, item in enumerate(instruments_data):
        current_row = start_data_row + idx
        ws.row_dimensions[current_row].height = 24
        row_fill = fill_zebra if idx % 2 == 1 else fill_white
        
        rec = item.get("recommendation", "Service")
        status_text = "ACTION REQ." if rec == "Replace" else "PASSED / OK"
        rec_font = Font(name="Segoe UI", size=9, bold=True, color="DC2626" if rec == "Replace" else "16A34A")

        row_data = [
            (idx + 1, align_center, font_data),
            (item.get("art_no", ""), align_center, font_bold),
            (item.get("name", ""), align_left, font_data),
            (item.get("damage", ""), align_left, font_data),
            (rec.upper(), align_center, rec_font),
            (status_text, align_center, font_data)
        ]

        for c_idx, (val, align, font_style) in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=c_idx, value=val)
            cell.font = font_style
            cell.fill = row_fill
            cell.alignment = align
            cell.border = cell_border

    col_widths = {1: 6, 2: 18, 3: 38, 4: 42, 5: 20, 6: 15}
    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

if 'num_instruments' not in st.session_state:
    st.session_state.num_instruments = 1

if 'pdf_generated' not in st.session_state:
    st.session_state.pdf_generated = False
if 'last_pdf_bytes' not in st.session_state:
    st.session_state.last_pdf_bytes = None
if 'last_excel_bytes' not in st.session_state:
    st.session_state.last_excel_bytes = None
if 'last_report_no' not in st.session_state:
    st.session_state.last_report_no = ""
if 'meta_payload' not in st.session_state:
    st.session_state.meta_payload = None
if 'instruments_payload' not in st.session_state:
    st.session_state.instruments_payload = None

def update_desc_callback(idx):
    sel_art = st.session_state.get(f"s_art_{idx}")
    if sel_art and sel_art in catalog_dict:
        st.session_state[f"name_{idx}"] = catalog_dict[sel_art]

# ==========================================
# 3. UI HEADER & SIDEBAR
# ==========================================
st.markdown(f"""
    <div class="brand-header">
        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 8px;">
            <div style="background: white; padding: 6px 12px; border-radius: 8px; display: inline-flex; align-items: center;">
                <img src="{LOGO_SRC}" alt="Biomed Logo" style="height: 28px; max-width: 100px; object-fit: contain;" />
            </div>
            <div class="status-badge">
                <span style="width: 8px; height: 8px; background-color: #4ADE80; border-radius: 50%; display: inline-block;"></span>
                SYSTEM ACTIVE
            </div>
        </div>
        <div>
            <h1>BIOMED INTERNATIONAL (PVT) LTD</h1>
            <p>AESCULAP DIVISION — EXECUTIVE TECHNICAL INSPECTION PORTAL</p>
        </div>
    </div>
""", unsafe_allow_html=True)

st.sidebar.markdown("### 📋 Inspection Context")
hospital_sel = st.sidebar.selectbox("Customer / Hospital", options=SL_HOSPITALS)
if hospital_sel == "Other (Type manually)":
    hospital_name = st.sidebar.text_input("Enter Hospital Name Manually")
elif hospital_sel == "-- Select Hospital / Institute --":
    hospital_name = ""
else:
    hospital_name = hospital_sel

date_val = st.sidebar.date_input("Inspection Date", value=datetime.date.today())
engineer_val = st.sidebar.text_input("Engineer / Inspector Name")
report_no_val = st.sidebar.text_input("Report Reference No.")
dept_val = st.sidebar.text_input("Department", value="Theatre / Laparoscopy")
remarks_val = st.sidebar.text_area("General Remarks & Inspection Notes", value="All above instruments require official technical evaluation and preventive maintenance as detailed.", height=90)

st.markdown("<div class='section-title'>🔬 Surgical Instruments Register</div>", unsafe_allow_html=True)

instruments_data = []

# ==========================================
# 4. INSTRUMENTS INPUT LOOP
# ==========================================
for i in range(st.session_state.num_instruments):
    st.markdown(f"<div class='instrument-card'><b>🔪 Instrument Entry #{i+1}</b>", unsafe_allow_html=True)
    
    inst_item = {}
    col_img, col_info = st.columns([1, 2])
    
    with col_img:
        inst_item["image"] = st.file_uploader(f"📷 Photo #{i+1}", type=["jpg", "png", "jpeg"], key=f"uploader_{i}")
        if inst_item["image"]:
            enhanced_preview = process_and_compress_image(inst_item["image"])
            st.image(enhanced_preview, caption="✨ Detail Enhanced Preview", use_container_width=True)
            
    with col_info:
        is_custom = st.checkbox("✍️ Custom Article No", key=f"custom_chk_{i}")
        if is_custom:
            art_no = st.text_input(f"Article No #{i+1}", key=f"c_art_{i}")
            inst_name = st.text_input(f"Instrument Description #{i+1}", key=f"name_{i}")
        else:
            art_no = st.selectbox(f"Search Master Catalog #{i+1}", options=[""] + article_options, key=f"s_art_{i}", on_change=update_desc_callback, args=(i,))
            inst_name = st.text_input(f"Instrument Description #{i+1}", key=f"name_{i}")
            
        inst_item["art_no"] = art_no
        inst_item["name"] = inst_name
        inst_item["sr_no"] = st.text_input(f"Machine Compatible / SR Number #{i+1}", key=f"sr_{i}")
        
        if inst_item["image"] and GEMINI_API_KEY:
            if st.button(f"✨ AI Auto-Detect Damage #{i+1}", key=f"ai_btn_{i}"):
                with st.spinner("Analyzing with AI..."):
                    ai_dam, ai_rec = analyze_damage_with_ai(inst_item["image"], inst_item["name"])
                    st.session_state[f"dam_{i}"] = ai_dam
                    st.session_state[f"rec_{i}"] = ai_rec
                    st.rerun()

        selected_preset = st.selectbox(f"💡 Technical Fault Presets #{i+1}", options=DAMAGE_SUGGESTIONS, key=f"preset_{i}")
        
        if selected_preset and not selected_preset.startswith("--"):
            curr = st.session_state.get(f"dam_{i}", "")
            if selected_preset not in curr:
                st.session_state[f"dam_{i}"] = f"{curr}\n{selected_preset}".strip() if curr else selected_preset

        inst_item["damage"] = st.text_area(f"Damage Details #{i+1}", key=f"dam_{i}", height=70)
        
        rec_opts = ["Replace", "Service", "Repair", "Upgrade / New System Required", "OK"]
        inst_item["recommendation"] = st.selectbox(f"Recommendation #{i+1}", options=rec_opts, key=f"rec_{i}")
        
    instruments_data.append(inst_item)
    st.markdown("</div>", unsafe_allow_html=True)

col_add, col_rem = st.columns(2)
with col_add:
    if st.button("➕ Add Another Instrument"):
        st.session_state.num_instruments += 1
        st.rerun()
with col_rem:
    if st.session_state.num_instruments > 1:
        if st.button("🗑️ Remove Last Instrument"):
            st.session_state.num_instruments -= 1
            st.rerun()

st.markdown("<br>", unsafe_allow_html=True)

# ==========================================
# 5. HIGH-END EXECUTIVE PDF GENERATION
# ==========================================
if st.button("📄 Build Executive PDF Report", type="primary", use_container_width=True):
    with st.spinner("Generating PDF Report..."):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=20,
            leftMargin=20,
            topMargin=20,
            bottomMargin=20
        )
        story, styles = [], getSampleStyleSheet()
        temp_files = []

        PRIMARY_NAVY = colors.HexColor("#0F172A")
        SECONDARY_SLATE = colors.HexColor("#334155")
        LIGHT_BG = colors.HexColor("#F8FAFC")
        BORDER_GRAY = colors.HexColor("#E2E8F0")

        title_style = ParagraphStyle('DocTitle', parent=styles['Heading1'], fontSize=12.5, leading=15, textColor=PRIMARY_NAVY, fontName="Helvetica-Bold")
        sub_style = ParagraphStyle('DocSub', parent=styles['Normal'], fontSize=8.5, leading=11, textColor=colors.HexColor("#64748B"))
        meta_label = ParagraphStyle('MetaLabel', parent=styles['Normal'], fontSize=8.5, leading=11, textColor=PRIMARY_NAVY, fontName="Helvetica-Bold")
        meta_val = ParagraphStyle('MetaVal', parent=styles['Normal'], fontSize=8.5, leading=11, textColor=SECONDARY_SLATE)
        cell_style = ParagraphStyle('TableCell', parent=styles['Normal'], fontSize=8.5, leading=11.5, textColor=SECONDARY_SLATE)
        cell_center = ParagraphStyle('TableCellCenter', parent=cell_style, alignment=1)
        th_style = ParagraphStyle('TH', parent=cell_style, fontSize=8.0, leading=10, textColor=colors.white, fontName="Helvetica-Bold", alignment=1)

        logo_img = RLImage("bmi_logo.png", width=75, height=32) if os.path.exists("bmi_logo.png") else Paragraph("<b>BIOMED</b>", title_style)
        
        comp_details = [
            Paragraph("BIOMED INTERNATIONAL (PVT) LTD", title_style),
            Paragraph("AESCULAP DIVISION | Colombo 03, Sri Lanka", sub_style)
        ]
        
        rep_title = [
            Paragraph("TECHNICAL INSPECTION REPORT", ParagraphStyle('RTitle', parent=title_style, alignment=2)),
            Paragraph("LAPAROSCOPY SYSTEM DIAGNOSTICS", ParagraphStyle('RSub', parent=sub_style, alignment=2))
        ]

        t_header = Table([[logo_img, comp_details, rep_title]], colWidths=[80, 260, 215])
        t_header.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ]))
        story.append(t_header)
        story.append(HRFlowable(width="100%", thickness=1.5, color=PRIMARY_NAVY, spaceBefore=0, spaceAfter=8))

        disp_hospital = hospital_name if hospital_name else "N/A"
        disp_engineer = engineer_val.strip() if engineer_val.strip() else "Biomed Technical Team"
        disp_rep_no = report_no_val.strip() if report_no_val.strip() else "N/A"
        date_str = date_val.strftime("%d %B %Y")

        meta_data = [
            [Paragraph("Customer / Hospital:", meta_label), Paragraph(disp_hospital, meta_val), Paragraph("Brand / System:", meta_label), Paragraph("Aesculap Laparoscopy", meta_val)],
            [Paragraph("Inspection Date:", meta_label), Paragraph(date_str, meta_val), Paragraph("Department:", meta_label), Paragraph(dept_val, meta_val)],
            [Paragraph("Engineer Name:", meta_label), Paragraph(disp_engineer, meta_val), Paragraph("Report Ref No:", meta_label), Paragraph(disp_rep_no, meta_val)],
        ]
        t_meta = Table(meta_data, colWidths=[95, 182, 95, 183])
        t_meta.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), LIGHT_BG),
            ('BOX', (0,0), (-1,-1), 0.5, BORDER_GRAY),
            ('INNERGRID', (0,0), (-1,-1), 0.5, BORDER_GRAY),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('PADDING', (0,0), (-1,-1), 5),
        ]))
        story.append(t_meta)
        story.append(Spacer(1, 10))

        table_data = [[
            Paragraph("#", th_style),
            Paragraph("INSPECTION PHOTO", th_style),
            Paragraph("ARTICLE NO", th_style),
            Paragraph("INSTRUMENT NAME", th_style),
            Paragraph("DETAILS OF DAMAGE / DEFECT", th_style),
            Paragraph("RECOMMENDATION", th_style)
        ]]

        for idx, item in enumerate(instruments_data):
            img_cell = Paragraph("No Image Attached", cell_center)
            if item["image"]:
                t_path = f"temp_p_{idx}.jpg"
                p_img = process_and_compress_image(item["image"], max_size=(1000, 1000))
                p_img.save(t_path, "JPEG", quality=95)
                
                img_cell = RLImage(t_path, width=115, height=105)
                temp_files.append(t_path)

            rec_text = item["recommendation"]
            rec_color = "#DC2626" if rec_text == "Replace" else ("#D97706" if rec_text in ["Service", "Repair"] else "#16A34A")

            table_data.append([
                Paragraph(str(idx + 1), cell_center),
                img_cell,
                Paragraph(f"<b>{item['art_no']}</b>", cell_center),
                Paragraph(f"<b>{item['name']}</b>", cell_style),
                Paragraph(item["damage"].replace("\n", "<br/>"), cell_style),
                Paragraph(f"<b><font color='{rec_color}'>{rec_text.upper()}</font></b>", cell_center)
            ])

        t_main = Table(table_data, colWidths=[18, 122, 65, 105, 155, 90])
        t_main.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), PRIMARY_NAVY),
            ('GRID', (0,0), (-1,-1), 0.5, BORDER_GRAY),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('LEFTPADDING', (0,0), (-1,-1), 4),
            ('RIGHTPADDING', (0,0), (-1,-1), 4),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, LIGHT_BG])
        ]))
        story.append(t_main)
        story.append(Spacer(1, 10))

        remarks_html = f"<b><font color='{PRIMARY_NAVY.hexval()}'>General Technical Remarks:</font></b><br/>{remarks_val.replace('\n', '<br/>')}"
        t_rem = Table([[Paragraph(remarks_html, cell_style)]], colWidths=[555])
        t_rem.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), LIGHT_BG),
            ('BOX', (0,0), (-1,-1), 0.5, BORDER_GRAY),
            ('PADDING', (0,0), (-1,-1), 6),
        ]))
        story.append(t_rem)
        story.append(Spacer(1, 15))

        sig_title_style = ParagraphStyle('SigTitle', parent=styles['Normal'], fontSize=8.5, leading=11, textColor=PRIMARY_NAVY, fontName="Helvetica-Bold")
        sig_text_style = ParagraphStyle('SigText', parent=styles['Normal'], fontSize=8.0, leading=10, textColor=SECONDARY_SLATE)

        sig_data = [
            [Paragraph("<b>Inspected & Prepared By:</b>", sig_title_style), Paragraph("<b>Customer Acknowledgment / Hospital Stamp:</b>", sig_title_style)],
            [Spacer(1, 22), Spacer(1, 22)],
            [Paragraph(f"........................................................<br/><b>Service Engineer:</b> {disp_engineer}<br/>Biomed International (Pvt) Ltd", sig_text_style),
             Paragraph("........................................................<br/><b>Authorized Signature & Stamp</b><br/>Hospital / Theatre Unit", sig_text_style)]
        ]

        t_sig = Table(sig_data, colWidths=[275, 280])
        t_sig.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ]))
        story.append(t_sig)

        doc.build(story)
        st.session_state.last_pdf_bytes = buffer.getvalue()

        for tf in temp_files:
            if os.path.exists(tf):
                os.remove(tf)

        st.session_state.meta_payload = {
            "report_no": disp_rep_no,
            "date": date_str,
            "hospital": disp_hospital,
            "engineer": disp_engineer
        }
        st.session_state.instruments_payload = instruments_data
        
        st.session_state.last_excel_bytes = generate_professional_excel(
            instruments_data=instruments_data,
            hospital_name=disp_hospital,
            engineer_name=disp_engineer,
            report_no=disp_rep_no,
            date_str=date_str
        )
        st.session_state.last_report_no = disp_rep_no
        st.session_state.pdf_generated = True

# ==========================================
# 6. DOWNLOADS & MANUAL GOOGLE SHEET SYNC
# ==========================================
if st.session_state.pdf_generated and st.session_state.last_pdf_bytes:
    st.success("✅ Executive PDF Report Ready!")
    
    col_dl1, col_dl2 = st.columns(2)
    with col_dl1:
        st.download_button(
            "📥 Download Executive PDF Report",
            data=st.session_state.last_pdf_bytes,
            file_name=f"Executive_Lap_Report_{st.session_state.last_report_no}.pdf",
            mime="application/pdf",
            use_container_width=True
        )
    with col_dl2:
        st.download_button(
            label="📊 Download Excel Technical Summary",
            data=st.session_state.last_excel_bytes,
            file_name=f"Lap_Report_Summary_{st.session_state.last_report_no}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("<div class='section-title'>📊 Manual Data Sync to Cloud</div>", unsafe_allow_html=True)
    
    if st.button("🔄 Sync Summary to Google Sheet Now", type="secondary", use_container_width=True):
        if st.session_state.get("instruments_payload") and st.session_state.get("meta_payload"):
            with st.spinner("Uploading items to Google Sheet..."):
                synced, err_msg = sync_to_google_sheet(
                    st.session_state.instruments_payload, 
                    st.session_state.meta_payload
                )
                if synced:
                    st.success("✅ All Instrument details successfully synced to Google Sheet!")
                else:
                    st.error(f"❌ Connection Error Details: {err_msg}")
        else:
            st.warning("⚠️ Please generate the PDF Report first before syncing.")
